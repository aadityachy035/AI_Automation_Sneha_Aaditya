"""
parse_catalogue.py
------------------
Reads the Tata Skylark IVI Message Catalogue Excel file (Main sheet)
and outputs a structured JSON file with the following fields per message:

  messagename  – "Message Name" value; prefix "CANRXFrame" is stripped if present
  isCAN        – True if Source Component is "CANNetwork", else False
  service id   – SOME/IP Service ID (decimal integer)
  method id    – SOME/IP Method ID  (decimal integer)
  direction    – "tx" if SOME/IP MSG Type is 0x01, "rx" if 0x02, else "unknown"

SETUP (run once in terminal):
  pip install openpyxl

USAGE:
  Place this script in the same folder as the .xlsx file, then run:
    python parse_catalogue.py

  Or pass paths explicitly:
    python parse_catalogue.py path/to/catalogue.xlsx path/to/output.json

OUTPUT:
  skylark_messages.json  (created in the same folder as the script)
"""

import sys
import json
import re
import os
import openpyxl


# ── Column indices (0-based) in the Main sheet ────────────────────────────────
COL_MESSAGE_NAME      = 14   # "Message Name"
COL_SOURCE_COMPONENT  = 17   # "Source Component"
COL_SERVICE_ID        = 24   # "SOME/IP Service ID"
COL_METHOD_ID         = 25   # "SOME/IP Meth. ID"
COL_MSG_TYPE          = 26   # "SOME/IP MSG Type"

# Prefix to strip from message names (case-insensitive match)
CANRXFRAME_PREFIX = "CANRXFrame"


def clean_message_name(raw) -> str:
    """
    Strip leading/trailing whitespace and non-breaking spaces.
    If the name starts with 'CANRXFrame' (case-insensitive), remove that
    prefix and return the remainder.
    """
    name = str(raw).replace("\xa0", "").strip()
    if name.upper().startswith(CANRXFRAME_PREFIX.upper()):
        name = name[len(CANRXFRAME_PREFIX):]   # remove prefix, keep the rest
    return name


def normalise_msg_type(raw) -> str:
    """
    Normalise messy MSG Type values like '\\xa00x01', ' 0x02', '2', etc.
    Returns a lowercase hex string like '0x01' / '0x02'.
    """
    if raw is None:
        return ""
    s = str(raw).strip().replace("\xa0", "").strip().lower()
    if re.match(r"^0x[0-9a-f]+$", s):
        return s
    if re.match(r"^\d+$", s):
        return hex(int(s))
    return s


def to_decimal(value):
    """
    Return the integer value of a service/method ID, or None if unparseable.
    Non-breaking spaces and plain whitespace are stripped before conversion.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).replace("\xa0", "").strip()
    try:
        # Handle hex strings that may appear in the sheet
        if s.lower().startswith("0x"):
            return int(s, 16)
        return int(s)
    except (ValueError, TypeError):
        return None


def parse_direction(msg_type_raw) -> str:
    t = normalise_msg_type(msg_type_raw)
    if t == "0x01":
        return "tx"
    if t == "0x02":
        return "rx"
    return "unknown"


def parse_catalogue(input_path: str, output_path: str) -> None:
    print(f"[+] Opening workbook : {input_path}")
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    if "Main" not in wb.sheetnames:
        raise ValueError("Sheet 'Main' not found in the workbook.")

    ws = wb["Main"]
    rows = list(ws.iter_rows(values_only=True))
    print(f"[+] Total rows (incl. header) : {len(rows)}")

    data_rows = rows[1:]   # skip header row
    results   = []
    skipped   = 0

    for row in data_rows:
        message_name     = row[COL_MESSAGE_NAME]
        source_component = row[COL_SOURCE_COMPONENT]
        service_id_raw   = row[COL_SERVICE_ID]
        method_id_raw    = row[COL_METHOD_ID]
        msg_type_raw     = row[COL_MSG_TYPE]

        # Skip rows with no meaningful data
        if message_name is None and service_id_raw is None:
            skipped += 1
            continue

        cleaned_name = clean_message_name(message_name) if message_name is not None else ""
        if not cleaned_name:
            skipped += 1
            continue

        # isCAN: True when Source Component is exactly "CANNetwork"
        src_comp_str = str(source_component).replace("\xa0", "").strip() \
                       if source_component is not None else ""
        is_can = src_comp_str == "CANNetwork"

        entry = {
            "messagename": cleaned_name,
            "isCAN":       is_can,
            "service id":  to_decimal(service_id_raw),
            "method id":   to_decimal(method_id_raw),
            "direction":   parse_direction(msg_type_raw),
        }
        results.append(entry)

    wb.close()

    print(f"[+] Parsed  : {len(results)} messages")
    print(f"[+] Skipped : {skipped} empty/blank rows")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"[+] Output  : {output_path}")


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Default: look for the xlsx next to this script
    script_dir     = os.path.dirname(os.path.abspath(__file__))
    default_input  = os.path.join(script_dir, "2025_TATA_SKYLARK_IVI_Program_Message_Catalogue_V1.2.1_ICE.xlsx")
    default_output = os.path.join(script_dir, "skylark_messages.json")

    input_file  = sys.argv[1] if len(sys.argv) > 1 else default_input
    output_file = sys.argv[2] if len(sys.argv) > 2 else default_output

    parse_catalogue(input_file, output_file)
