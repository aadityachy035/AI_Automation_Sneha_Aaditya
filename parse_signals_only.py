"""
parse_signals_only.py
---------------------
Reads the Tata Skylark IVI Message Catalogue Excel file (Main sheet)
and outputs a JSON with only messageName and signals array.
Messages where Source Component = CANNetwork are excluded.

Output format:
[
  {
    "messageName": "SOCSession",
    "signals": ["socSession[0]", "socSession[1]"]
  },
  ...
]

USAGE:
  C:\\Users\\achoudh4\\AppData\\Local\\Programs\\Python\\Python311\\python.exe parse_signals_only.py
"""

import sys
import json
import re
import os
import openpyxl


COL_MESSAGE_NAME      = 14   # "Message Name"
COL_SOURCE_COMPONENT  = 17   # "Source Component"
COL_SIGNAL_NAME       = 28   # "Signal Name"

CANRXFRAME_PREFIX = "CANRXFrame"


def clean_message_name(raw) -> str:
    name = str(raw).replace("\xa0", "").strip()
    if name.upper().startswith(CANRXFRAME_PREFIX.upper()):
        name = name[len(CANRXFRAME_PREFIX):]
    return name


def expand_signal(raw_signal: str) -> list:
    """
    signalName[3]  ->  signalName[0], signalName[1], signalName[2]
    plain name     ->  kept as-is
    multi-line cell -> each line processed separately
    """
    signals = []
    parts = re.split(r'[\r\n]+', raw_signal)
    for part in parts:
        part = part.replace("\xa0", "").strip()
        if not part:
            continue
        m = re.match(r'^(.+?)\[(\d+)\]$', part)
        if m:
            base  = m.group(1).strip()
            count = int(m.group(2))
            for i in range(count):
                signals.append(f"{base}[{i}]")
        else:
            signals.append(part)
    return signals


def parse_signals(input_path: str, output_path: str) -> None:
    print(f"[+] Opening workbook : {input_path}")
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    if "Main" not in wb.sheetnames:
        raise ValueError("Sheet 'Main' not found in the workbook.")

    ws   = wb["Main"]
    rows = list(ws.iter_rows(values_only=True))
    print(f"[+] Total rows (incl. header) : {len(rows)}")

    results       = []
    current_entry = None   # tracks the active message; None = skip continuation rows too
    skipped_can   = 0

    for row in rows[1:]:   # skip header
        message_name = row[COL_MESSAGE_NAME]
        signal_raw   = row[COL_SIGNAL_NAME]

        # ── New message row (Message Name column is filled) ──────────────────
        if message_name is not None:
            cleaned_name = clean_message_name(message_name)
            if not cleaned_name:
                current_entry = None
                continue

            # Skip messages whose Source Component is CANNetwork
            src_comp = row[COL_SOURCE_COMPONENT]
            src_comp_str = str(src_comp).replace("\xa0", "").strip() if src_comp is not None else ""
            if src_comp_str == "CANNetwork":
                current_entry = None   # also blocks continuation rows below
                skipped_can += 1
                continue

            current_entry = {
                "messageName": cleaned_name,
                "signals":     [],
            }
            results.append(current_entry)

        # ── Signal row: add to active message (current_entry = None means skip) ──
        if current_entry is not None and signal_raw is not None:
            sig_str = str(signal_raw).replace("\xa0", "").strip()
            if sig_str:
                current_entry["signals"].extend(expand_signal(sig_str))

    wb.close()

    print(f"[+] Parsed   : {len(results)} messages")
    print(f"[+] Skipped  : {skipped_can} CANNetwork messages")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"[+] Output   : {output_path}")


if __name__ == "__main__":
    script_dir     = os.path.dirname(os.path.abspath(__file__))
    default_input  = os.path.join(script_dir, "2025_TATA_SKYLARK_IVI_Program_Message_Catalogue_V1.2.1_ICE.xlsx")
    default_output = os.path.join(script_dir, "skylark_signals.json")

    input_file  = sys.argv[1] if len(sys.argv) > 1 else default_input
    output_file = sys.argv[2] if len(sys.argv) > 2 else default_output

    parse_signals(input_file, output_file)